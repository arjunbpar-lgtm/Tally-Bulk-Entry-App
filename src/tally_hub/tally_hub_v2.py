import argparse
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

try:
    from . import monthly_client_tally_converter as base
    from .defaults import (
        DEFAULT_BATCH,
        DEFAULT_COMPANY_PROFILE_NAME,
        DEFAULT_FALLBACK_UNIT,
        DEFAULT_GODOWN,
        DEFAULT_PURCHASE_LEDGER,
        DEFAULT_PURCHASE_LOCAL_LEDGER,
        DEFAULT_PURCHASE_PARTY_PARENT,
        DEFAULT_SALES_LEDGER,
        DEFAULT_SALES_PARTY_PARENT,
        DEFAULT_STOCK_PARENT,
        TALLY_HUB_APP_TITLE,
        TALLY_HUB_STATE_FILE,
    )
except ImportError:
    import monthly_client_tally_converter as base
    from defaults import (
        DEFAULT_BATCH,
        DEFAULT_COMPANY_PROFILE_NAME,
        DEFAULT_FALLBACK_UNIT,
        DEFAULT_GODOWN,
        DEFAULT_PURCHASE_LEDGER,
        DEFAULT_PURCHASE_LOCAL_LEDGER,
        DEFAULT_PURCHASE_PARTY_PARENT,
        DEFAULT_SALES_LEDGER,
        DEFAULT_SALES_PARTY_PARENT,
        DEFAULT_STOCK_PARENT,
        TALLY_HUB_APP_TITLE,
        TALLY_HUB_STATE_FILE,
    )


APP_TITLE = TALLY_HUB_APP_TITLE
STATE_FILE = TALLY_HUB_STATE_FILE
COMPANY_PROFILE_NAME = DEFAULT_COMPANY_PROFILE_NAME


def _sanitize_output_folder_name(raw_name: object) -> str:
    folder_name = base.sanitize_xml_text(raw_name)
    folder_name = re.sub(r'[<>:"/\\|?*]+', " ", folder_name)
    folder_name = re.sub(r"\s+", " ", folder_name).strip().rstrip(".")
    return folder_name


@dataclass
class HubConfig:
    company_name: str = ""
    sales_ledger_name: str = DEFAULT_SALES_LEDGER
    purchase_ledger_name: str = DEFAULT_PURCHASE_LEDGER
    purchase_local_ledger_name: str = DEFAULT_PURCHASE_LOCAL_LEDGER
    sales_party_parent: str = DEFAULT_SALES_PARTY_PARENT
    purchase_party_parent: str = DEFAULT_PURCHASE_PARTY_PARENT
    purchase_party_name_override: str = ""
    stock_parent_group: str = DEFAULT_STOCK_PARENT
    godown_name: str = DEFAULT_GODOWN
    batch_name: str = DEFAULT_BATCH
    fallback_unit: str = DEFAULT_FALLBACK_UNIT
    output_dir: str = "."


@dataclass
class PurchaseVoucher:
    voucher_no: str
    date: str
    party_name: str
    supplier_state_code: str
    purchase_ledger_name: str
    total_amount: Decimal
    lines: List[base.SalesLine]


def _extract_supplier_state_code(gst_no: object, state_code: object = "") -> str:
    gst_text = base.sanitize_xml_text(gst_no)
    match = re.match(r"\d{2}", gst_text)
    if match:
        return match.group(0)
    state_text = base.sanitize_xml_text(state_code)
    match = re.match(r"\d{1,2}", state_text)
    if match:
        return match.group(0).zfill(2)
    return ""


def _purchase_ledger_for_state_code(supplier_state_code: str, config: HubConfig) -> str:
    if supplier_state_code == "32":
        return base.sanitize_xml_text(config.purchase_local_ledger_name or DEFAULT_PURCHASE_LOCAL_LEDGER)
    return base.sanitize_xml_text(config.purchase_ledger_name or DEFAULT_PURCHASE_LEDGER)


def _read_state(state_path: Path) -> Dict[str, List[str]]:
    if not state_path.exists():
        return {"known_ledgers": [], "known_stock_items": []}
    try:
        data = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return {"known_ledgers": [], "known_stock_items": []}
    if not isinstance(data, dict):
        return {"known_ledgers": [], "known_stock_items": []}
    known_ledgers = data.get("known_ledgers", [])
    known_stock_items = data.get("known_stock_items", [])
    if not isinstance(known_ledgers, list) or not isinstance(known_stock_items, list):
        return {"known_ledgers": [], "known_stock_items": []}
    return {
        "known_ledgers": [base.sanitize_xml_text(x) for x in known_ledgers if base.sanitize_xml_text(x)],
        "known_stock_items": [base.sanitize_xml_text(x) for x in known_stock_items if base.sanitize_xml_text(x)],
    }


def _write_state(state_path: Path, known_ledgers: Iterable[str], known_stock_items: Iterable[str]) -> None:
    payload = {
        "known_ledgers": sorted({base.sanitize_xml_text(x) for x in known_ledgers if base.sanitize_xml_text(x)}),
        "known_stock_items": sorted({base.sanitize_xml_text(x) for x in known_stock_items if base.sanitize_xml_text(x)}),
    }
    state_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def parse_purchase_report_excel(
    excel_path: str,
    master_snapshot: Optional[base.MasterSnapshot] = None,
) -> Tuple[List[PurchaseVoucher], List[str], List[Tuple[str, str, str]], List[str]]:
    master_snapshot = master_snapshot or base.MasterSnapshot()
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = None
    for row_idx in range(1, min(worksheet.max_row, 40) + 1):
        row_values = {base.sanitize_xml_text(worksheet.cell(row_idx, col).value) for col in range(1, worksheet.max_column + 1)}
        if {"Date", "VCh.No", "AcName", "ItemName", "Qty", "Unit"}.issubset(row_values):
            header_row = row_idx
            break
    if header_row is None:
        raise ValueError("Could not detect purchase header row in the Excel report.")

    header_map: Dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        header_name = base.sanitize_xml_text(worksheet.cell(header_row, col).value)
        if header_name:
            header_map[header_name] = col

    required_headers = ["Date", "VCh.No", "AcName", "ItemName", "Qty", "Unit"]
    missing_headers = [name for name in required_headers if name not in header_map]
    if missing_headers:
        raise ValueError("Missing required purchase columns: " + ", ".join(missing_headers))

    grouped_rows: Dict[Tuple[str, str, str, str], List[base.SalesLine]] = {}
    skipped_rows: List[str] = []
    item_resolution_cache: Dict[str, str] = {}
    missing_stock_candidates: Dict[str, Tuple[str, str, str]] = {}

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        sl_no = base.sanitize_xml_text(
            worksheet.cell(row_idx, header_map.get("SlNo", 0)).value if "SlNo" in header_map else ""
        )
        voucher_no = base.sanitize_xml_text(worksheet.cell(row_idx, header_map["VCh.No"]).value)
        raw_date = worksheet.cell(row_idx, header_map["Date"]).value
        party_name = base.sanitize_xml_text(worksheet.cell(row_idx, header_map["AcName"]).value)
        gst_no = base.sanitize_xml_text(worksheet.cell(row_idx, header_map.get("GST No", 0)).value if "GST No" in header_map else "")
        state_code = base.sanitize_xml_text(worksheet.cell(row_idx, header_map.get("State Code", 0)).value if "State Code" in header_map else "")
        supplier_state_code = _extract_supplier_state_code(gst_no, state_code)
        if "Party Ledger" in header_map:
            override_name = base.sanitize_xml_text(worksheet.cell(row_idx, header_map["Party Ledger"]).value)
            if override_name:
                party_name = override_name
        item_name = base.sanitize_xml_text(worksheet.cell(row_idx, header_map["ItemName"]).value)
        hsn = base.sanitize_xml_text(worksheet.cell(row_idx, header_map.get("HSN", 0)).value if "HSN" in header_map else "")
        qty = base.parse_amount(worksheet.cell(row_idx, header_map["Qty"]).value)
        unit_name = base.sanitize_xml_text(worksheet.cell(row_idx, header_map["Unit"]).value)

        amount = None
        if "Total" in header_map:
            amount = base.parse_amount(worksheet.cell(row_idx, header_map["Total"]).value)
        if amount is None and "Amount" in header_map:
            amount = base.parse_amount(worksheet.cell(row_idx, header_map["Amount"]).value)

        if sl_no and not sl_no.isdigit():
            continue
        if not any([voucher_no, raw_date, party_name, item_name, qty, amount]):
            continue
        if item_name.lower() == "total":
            continue

        normalized_date = base.normalize_date(raw_date)
        row_errors = []
        if not voucher_no:
            row_errors.append("voucher no")
        if not normalized_date:
            row_errors.append("date")
        if not party_name:
            row_errors.append("supplier")
        if not item_name:
            row_errors.append("item")
        if qty is None:
            row_errors.append("qty")
        if amount is None:
            row_errors.append("amount")
        if row_errors:
            skipped_rows.append(f"Row {row_idx}: {', '.join(row_errors)}")
            continue

        resolved_item_name = item_resolution_cache.get(item_name)
        if not resolved_item_name:
            normalized_item = base.normalize_key(item_name)
            resolved_item_name = master_snapshot.stock_items.get(normalized_item, item_name)
            item_resolution_cache[item_name] = resolved_item_name
        if resolved_item_name == item_name and base.normalize_key(item_name) not in master_snapshot.stock_items:
            fallback_unit = unit_name or master_snapshot.units.get(
                base.normalize_key(DEFAULT_FALLBACK_UNIT),
                DEFAULT_FALLBACK_UNIT,
            )
            missing_stock_candidates[base.normalize_key(item_name)] = (item_name, hsn, fallback_unit)

        resolved_unit = unit_name or master_snapshot.units.get(
            base.normalize_key(DEFAULT_FALLBACK_UNIT),
            DEFAULT_FALLBACK_UNIT,
        )
        if unit_name and master_snapshot.units:
            resolved_unit = master_snapshot.units.get(base.normalize_key(unit_name), unit_name)

        key = (normalized_date, voucher_no, party_name, supplier_state_code)
        grouped_rows.setdefault(key, []).append(
            base.SalesLine(
                item_name=item_name,
                resolved_item_name=resolved_item_name,
                hsn=hsn,
                qty=abs(qty),
                unit_name=resolved_unit or DEFAULT_FALLBACK_UNIT,
                amount=abs(amount),
            )
        )

    vouchers: List[PurchaseVoucher] = []
    for (voucher_date, voucher_no, party_name, supplier_state_code), lines in sorted(grouped_rows.items(), key=lambda item: (item[0][0], item[0][1])):
        total_amount = sum((line.amount for line in lines), Decimal("0"))
        vouchers.append(
            PurchaseVoucher(
                voucher_no=voucher_no,
                date=voucher_date,
                party_name=party_name,
                supplier_state_code=supplier_state_code,
                purchase_ledger_name="",
                total_amount=total_amount,
                lines=lines,
            )
        )

    supplier_names = sorted({voucher.party_name for voucher in vouchers}, key=str.upper)
    missing_stock_items = [missing_stock_candidates[key] for key in sorted(missing_stock_candidates)]
    return vouchers, supplier_names, missing_stock_items, skipped_rows


def _add_common_empty_lists(parent: ET.Element, tag_names: Sequence[str]) -> None:
    for tag_name in tag_names:
        ET.SubElement(parent, tag_name)


def _build_purchase_inventory_line(
    parent: ET.Element,
    line: base.SalesLine,
    purchase_ledger_name: str,
    godown_name: str,
    batch_name: str,
) -> None:
    inventory = ET.SubElement(parent, "ALLINVENTORYENTRIES.LIST")
    ET.SubElement(inventory, "STOCKITEMNAME").text = base.sanitize_xml_text(line.resolved_item_name)
    ET.SubElement(inventory, "GSTSOURCETYPE").text = "Ledger"
    ET.SubElement(inventory, "GSTLEDGERSOURCE").text = base.sanitize_xml_text(purchase_ledger_name)
    ET.SubElement(inventory, "HSNSOURCETYPE").text = "Stock Item"
    ET.SubElement(inventory, "HSNITEMSOURCE").text = base.sanitize_xml_text(line.resolved_item_name)
    ET.SubElement(inventory, "GSTOVRDNTYPEOFSUPPLY").text = "Goods"
    ET.SubElement(inventory, "GSTRATEINFERAPPLICABILITY").text = "As per Masters/Company"
    if line.hsn:
        ET.SubElement(inventory, "GSTHSNNAME").text = base.sanitize_xml_text(line.hsn)
    ET.SubElement(inventory, "GSTHSNINFERAPPLICABILITY").text = "As per Masters/Company"
    ET.SubElement(inventory, "ISDEEMEDPOSITIVE").text = "Yes"

    rate_value = Decimal("0")
    if line.qty > 0:
        rate_value = (line.amount / line.qty).quantize(Decimal("0.01"))
    ET.SubElement(inventory, "RATE").text = f"{base.format_amount(rate_value)}/{base.sanitize_xml_text(line.unit_name)}"
    ET.SubElement(inventory, "AMOUNT").text = base.format_amount(-line.amount)
    ET.SubElement(inventory, "ACTUALQTY").text = base.format_qty(line.qty, line.unit_name)
    ET.SubElement(inventory, "BILLEDQTY").text = base.format_qty(line.qty, line.unit_name)

    batch = ET.SubElement(inventory, "BATCHALLOCATIONS.LIST")
    ET.SubElement(batch, "GODOWNNAME").text = base.sanitize_xml_text(godown_name)
    ET.SubElement(batch, "BATCHNAME").text = base.sanitize_xml_text(batch_name)
    ET.SubElement(batch, "AMOUNT").text = base.format_amount(-line.amount)
    ET.SubElement(batch, "ACTUALQTY").text = base.format_qty(line.qty, line.unit_name)
    ET.SubElement(batch, "BILLEDQTY").text = base.format_qty(line.qty, line.unit_name)
    _add_common_empty_lists(batch, ["ADDITIONALDETAILS.LIST", "VOUCHERCOMPONENTLIST.LIST"])

    accounting = ET.SubElement(inventory, "ACCOUNTINGALLOCATIONS.LIST")
    ET.SubElement(accounting, "LEDGERNAME").text = base.sanitize_xml_text(purchase_ledger_name)
    ET.SubElement(accounting, "ISDEEMEDPOSITIVE").text = "Yes"
    ET.SubElement(accounting, "LEDGERFROMITEM").text = "No"
    ET.SubElement(accounting, "ISPARTYLEDGER").text = "No"
    ET.SubElement(accounting, "AMOUNT").text = base.format_amount(-line.amount)
    _add_common_empty_lists(
        accounting,
        [
            "BANKALLOCATIONS.LIST",
            "BILLALLOCATIONS.LIST",
            "INTERESTCOLLECTION.LIST",
            "OLDAUDITENTRIES.LIST",
            "ACCOUNTAUDITENTRIES.LIST",
            "AUDITENTRIES.LIST",
            "INPUTCRALLOCS.LIST",
            "DUTYHEADDETAILS.LIST",
            "RATEDETAILS.LIST",
            "SUMMARYALLOCS.LIST",
            "TAXBILLALLOCATIONS.LIST",
            "TAXOBJECTALLOCATIONS.LIST",
            "VATSTATUTORYDETAILS.LIST",
        ],
    )

    _add_common_empty_lists(
        inventory,
        [
            "DUTYHEADDETAILS.LIST",
            "RATEDETAILS.LIST",
            "SUPPLEMENTARYDUTYHEADDETAILS.LIST",
            "TAXOBJECTALLOCATIONS.LIST",
            "REFVOUCHERDETAILS.LIST",
            "EXCISEALLOCATIONS.LIST",
            "EXPENSEALLOCATIONS.LIST",
        ],
    )


def build_purchase_vouchers_xml(
    vouchers: Sequence[PurchaseVoucher],
    company_name: str,
    purchase_ledger_name: str,
    godown_name: str,
    batch_name: str,
    output_path: str,
) -> str:
    root, request_data = base.new_envelope("Vouchers", company_name)

    for voucher_row in vouchers:
        party_name = base.sanitize_xml_text(voucher_row.party_name)
        invoice_no = base.sanitize_xml_text(voucher_row.voucher_no)
        voucher_purchase_ledger = base.sanitize_xml_text(voucher_row.purchase_ledger_name or purchase_ledger_name)
        tally_message = ET.SubElement(request_data, "TALLYMESSAGE", attrib={"xmlns:UDF": "TallyUDF"})
        voucher = ET.SubElement(
            tally_message,
            "VOUCHER",
            attrib={"VCHTYPE": "Purchase", "ACTION": "Create", "OBJVIEW": "Invoice Voucher View"},
        )
        ET.SubElement(voucher, "DATE").text = voucher_row.date
        ET.SubElement(voucher, "REFERENCEDATE").text = voucher_row.date
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = "Purchase"
        ET.SubElement(voucher, "REFERENCE").text = invoice_no
        ET.SubElement(voucher, "PARTYLEDGERNAME").text = party_name
        ET.SubElement(voucher, "BASICBASEPARTYNAME").text = party_name
        ET.SubElement(voucher, "PARTYNAME").text = party_name
        ET.SubElement(voucher, "VOUCHERNUMBER").text = invoice_no
        ET.SubElement(voucher, "PERSISTEDVIEW").text = "Invoice Voucher View"
        ET.SubElement(voucher, "ISINVOICE").text = "Yes"
        ET.SubElement(voucher, "OBJVIEW").text = "Invoice Voucher View"
        ET.SubElement(voucher, "EFFECTIVEDATE").text = voucher_row.date

        party_entry = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
        ET.SubElement(party_entry, "LEDGERNAME").text = party_name
        ET.SubElement(party_entry, "ISDEEMEDPOSITIVE").text = "No"
        ET.SubElement(party_entry, "LEDGERFROMITEM").text = "No"
        ET.SubElement(party_entry, "ISPARTYLEDGER").text = "Yes"
        ET.SubElement(party_entry, "ISLASTDEEMEDPOSITIVE").text = "No"
        ET.SubElement(party_entry, "AMOUNT").text = base.format_amount(voucher_row.total_amount)
        ET.SubElement(party_entry, "BANKALLOCATIONS.LIST")
        bill_allocation = ET.SubElement(party_entry, "BILLALLOCATIONS.LIST")
        ET.SubElement(bill_allocation, "NAME").text = invoice_no
        ET.SubElement(bill_allocation, "BILLTYPE").text = "New Ref"
        ET.SubElement(bill_allocation, "AMOUNT").text = base.format_amount(voucher_row.total_amount)
        _add_common_empty_lists(
            party_entry,
            [
                "INTERESTCOLLECTION.LIST",
                "OLDAUDITENTRIES.LIST",
                "ACCOUNTAUDITENTRIES.LIST",
                "AUDITENTRIES.LIST",
                "INPUTCRALLOCS.LIST",
                "DUTYHEADDETAILS.LIST",
                "RATEDETAILS.LIST",
                "SUMMARYALLOCS.LIST",
                "TAXBILLALLOCATIONS.LIST",
                "TAXOBJECTALLOCATIONS.LIST",
                "VATSTATUTORYDETAILS.LIST",
            ],
        )

        for line in voucher_row.lines:
            _build_purchase_inventory_line(
                parent=voucher,
                line=line,
                purchase_ledger_name=voucher_purchase_ledger,
                godown_name=godown_name,
                batch_name=batch_name,
            )

        _add_common_empty_lists(
            voucher,
            [
                "GST.LIST",
                "GSTBUYERADDRESS.LIST",
                "GSTCONSIGNEEADDRESS.LIST",
                "PAYROLLMODEOFPAYMENT.LIST",
                "ATTDRECORDS.LIST",
            ],
        )

    base.safe_indent(root)
    ET.ElementTree(root).write(output_path, encoding="utf-8", xml_declaration=True)
    return output_path


def convert_purchase_report(
    excel_path: str,
    master_xml_path: Optional[str],
    config: HubConfig,
    state_path: Optional[str] = None,
) -> base.ConversionResult:
    output_dir = Path(config.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    state_path_obj = Path(state_path or (output_dir / STATE_FILE))
    state = _read_state(state_path_obj)
    master_snapshot = base.load_master_snapshot(master_xml_path)

    if not config.company_name and master_snapshot.company_name:
        config.company_name = master_snapshot.company_name
    if master_snapshot.godowns and base.normalize_key(config.godown_name) in master_snapshot.godowns:
        config.godown_name = master_snapshot.godowns[base.normalize_key(config.godown_name)]
    if master_snapshot.units and base.normalize_key(config.fallback_unit) in master_snapshot.units:
        config.fallback_unit = master_snapshot.units[base.normalize_key(config.fallback_unit)]

    vouchers, supplier_names, missing_stock_items, skipped_rows = parse_purchase_report_excel(excel_path, master_snapshot)
    vouchers = [
        PurchaseVoucher(
            voucher_no=voucher.voucher_no,
            date=voucher.date,
            party_name=voucher.party_name,
            supplier_state_code=voucher.supplier_state_code,
            purchase_ledger_name=_purchase_ledger_for_state_code(voucher.supplier_state_code, config),
            total_amount=voucher.total_amount,
            lines=voucher.lines,
        )
        for voucher in vouchers
    ]

    if config.purchase_party_name_override:
        override_name = base.sanitize_xml_text(config.purchase_party_name_override)
        vouchers = [
            PurchaseVoucher(
                voucher_no=voucher.voucher_no,
                date=voucher.date,
                party_name=override_name,
                supplier_state_code=voucher.supplier_state_code,
                purchase_ledger_name=voucher.purchase_ledger_name,
                total_amount=voucher.total_amount,
                lines=voucher.lines,
            )
            for voucher in vouchers
        ]
        supplier_names = [override_name] if override_name else supplier_names

    known_ledgers = {base.sanitize_xml_text(name) for name in state["known_ledgers"]}
    known_ledgers.update(master_snapshot.ledgers.values())
    reserved_ledgers = {"cash"}
    suppliers_to_export = [
        name
        for name in supplier_names
        if base.normalize_key(name) not in reserved_ledgers
    ]

    output_files: List[str] = []

    supplier_xml_path = output_dir / "01_supplier_ledgers.xml"
    supplier_cfg = base.ConversionConfig(
        company_name=config.company_name,
        party_parent_group=config.purchase_party_parent,
    )
    output_files.append(base.build_ledger_masters_xml(suppliers_to_export, supplier_cfg, str(supplier_xml_path)))

    stock_xml_path = output_dir / "02_stock_masters_missing_purchase.xml"
    stock_cfg = base.ConversionConfig(
        company_name=config.company_name,
        stock_parent_group=config.stock_parent_group,
        fallback_unit=config.fallback_unit,
    )
    stock_output = base.build_stock_masters_xml(missing_stock_items, stock_cfg, str(stock_xml_path))
    if stock_output:
        output_files.append(stock_output)
    elif stock_xml_path.exists():
        stock_xml_path.unlink()

    purchase_xml_path = output_dir / "03_purchase_vouchers.xml"
    output_files.append(
        build_purchase_vouchers_xml(
            vouchers=vouchers,
            company_name=config.company_name,
            purchase_ledger_name=config.purchase_ledger_name,
            godown_name=config.godown_name,
            batch_name=config.batch_name,
            output_path=str(purchase_xml_path),
        )
    )

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "voucher_count": len(vouchers),
        "voucher_lines": sum(len(v.lines) for v in vouchers),
        "new_ledgers": list(suppliers_to_export),
        "purchase_ledgers": sorted({voucher.purchase_ledger_name for voucher in vouchers if voucher.purchase_ledger_name}),
        "missing_stock_items": [
            {"item_name": item_name, "hsn": hsn, "unit_name": unit_name}
            for item_name, hsn, unit_name in missing_stock_items
        ],
        "skipped_rows": list(skipped_rows),
    }
    summary_path = output_dir / "purchase_conversion_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    output_files.append(str(summary_path))

    bundle_path = output_dir / "purchase_import_bundle.zip"
    with zipfile.ZipFile(bundle_path, "w", compression=zipfile.ZIP_DEFLATED) as bundle:
        for file_path in output_files:
            bundle.write(file_path, arcname=Path(file_path).name)
    output_files.append(str(bundle_path))

    updated_ledgers = known_ledgers | {base.sanitize_xml_text(name) for name in supplier_names}
    updated_stock_items = {base.sanitize_xml_text(name) for name in state["known_stock_items"]}
    updated_stock_items.update(master_snapshot.stock_items.values())
    for voucher in vouchers:
        for line in voucher.lines:
            updated_stock_items.add(base.sanitize_xml_text(line.resolved_item_name))
    _write_state(state_path_obj, updated_ledgers, updated_stock_items)

    return base.ConversionResult(
        vouchers=vouchers,
        ledger_names=suppliers_to_export,
        missing_stock_items=missing_stock_items,
        output_files=output_files,
        skipped_rows=skipped_rows,
        summary_path=str(summary_path),
    )


def launch_gui() -> None:
    try:
        from PySide6.QtWidgets import (
            QApplication,
            QFileDialog,
            QFormLayout,
            QGridLayout,
            QGroupBox,
            QHBoxLayout,
            QInputDialog,
            QLabel,
            QLineEdit,
            QMainWindow,
            QMessageBox,
            QPushButton,
            QStackedWidget,
            QTextEdit,
            QVBoxLayout,
            QWidget,
        )
    except Exception as ex:
        raise RuntimeError("PySide6 is required for GUI mode.") from ex

    class HubWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle(APP_TITLE)
            self.resize(1120, 780)
            self.current_module = "Sales"
            self.report_excel_path = ""
            self.master_xml_path = ""
            self._build_ui()

        def _build_ui(self):
            root = QWidget()
            self.setCentralWidget(root)
            outer = QVBoxLayout(root)
            outer.setContentsMargins(16, 16, 16, 16)
            outer.setSpacing(10)

            self.stack = QStackedWidget()
            outer.addWidget(self.stack, 1)

            self.company_page = self._build_company_page()
            self.module_page = self._build_module_page()
            self.workspace_page = self._build_workspace_page()
            self.stack.addWidget(self.company_page)
            self.stack.addWidget(self.module_page)
            self.stack.addWidget(self.workspace_page)
            self.stack.setCurrentIndex(0)

        def _build_company_page(self):
            page = QWidget()
            layout = QVBoxLayout(page)
            title = QLabel("Select Company")
            title.setStyleSheet("font-size: 26px; font-weight: 700;")
            subtitle = QLabel("Choose the company profile to continue.")
            select_btn = QPushButton(COMPANY_PROFILE_NAME)
            select_btn.setMinimumHeight(90)
            select_btn.clicked.connect(lambda: self.stack.setCurrentIndex(1))
            layout.addWidget(title)
            layout.addWidget(subtitle)
            layout.addStretch()
            layout.addWidget(select_btn)
            layout.addStretch()
            return page

        def _build_module_page(self):
            page = QWidget()
            layout = QVBoxLayout(page)
            title = QLabel(COMPANY_PROFILE_NAME)
            title.setStyleSheet("font-size: 24px; font-weight: 700;")
            subtitle = QLabel("Select module")
            sales_btn = QPushButton("Sales Entry")
            sales_btn.setMinimumHeight(90)
            sales_btn.clicked.connect(lambda: self._open_module("Sales"))
            purchase_btn = QPushButton("Purchase Entry")
            purchase_btn.setMinimumHeight(90)
            purchase_btn.clicked.connect(lambda: self._open_module("Purchase"))
            back_btn = QPushButton("Back to Company")
            back_btn.clicked.connect(lambda: self.stack.setCurrentIndex(0))
            layout.addWidget(title)
            layout.addWidget(subtitle)
            layout.addWidget(sales_btn)
            layout.addWidget(purchase_btn)
            layout.addStretch()
            layout.addWidget(back_btn)
            return page

        def _build_workspace_page(self):
            page = QWidget()
            layout = QVBoxLayout(page)

            self.workspace_title = QLabel(f"{COMPANY_PROFILE_NAME} - Sales Entry")
            self.workspace_title.setStyleSheet("font-size: 20px; font-weight: 700;")
            layout.addWidget(self.workspace_title)

            files_box = QGroupBox("Source Files")
            files_grid = QGridLayout(files_box)
            self.report_path_input = QLineEdit()
            self.master_path_input = QLineEdit()
            self.report_path_input.setReadOnly(True)
            self.master_path_input.setReadOnly(True)
            load_report_btn = QPushButton("Load Report Excel")
            load_report_btn.clicked.connect(self._pick_report_excel)
            load_master_btn = QPushButton("Load Master XML")
            load_master_btn.clicked.connect(self._pick_master_xml)
            files_grid.addWidget(QLabel("Report file"), 0, 0)
            files_grid.addWidget(self.report_path_input, 0, 1)
            files_grid.addWidget(load_report_btn, 0, 2)
            files_grid.addWidget(QLabel("Existing Master XML"), 1, 0)
            files_grid.addWidget(self.master_path_input, 1, 1)
            files_grid.addWidget(load_master_btn, 1, 2)
            layout.addWidget(files_box)

            settings_box = QGroupBox("Import Settings")
            settings_form = QFormLayout(settings_box)
            self.company_input = QLineEdit()
            self.ledger_input = QLineEdit(DEFAULT_SALES_LEDGER)
            self.purchase_local_ledger_input = QLineEdit(DEFAULT_PURCHASE_LOCAL_LEDGER)
            self.party_parent_input = QLineEdit(DEFAULT_SALES_PARTY_PARENT)
            self.party_override_input = QLineEdit()
            self.stock_parent_input = QLineEdit(DEFAULT_STOCK_PARENT)
            self.godown_input = QLineEdit(DEFAULT_GODOWN)
            self.batch_input = QLineEdit(DEFAULT_BATCH)
            self.output_dir_input = QLineEdit(str(Path.cwd()))
            self.ledger_label = QLabel("Sales ledger")
            self.purchase_local_ledger_label = QLabel("Local purchase ledger")
            self.party_parent_label = QLabel("Party parent group")
            self.party_override_label = QLabel("Party ledger override")
            settings_form.addRow("Company name", self.company_input)
            settings_form.addRow(self.ledger_label, self.ledger_input)
            settings_form.addRow(self.purchase_local_ledger_label, self.purchase_local_ledger_input)
            self.purchase_local_ledger_label.hide()
            self.purchase_local_ledger_input.hide()
            settings_form.addRow(self.party_parent_label, self.party_parent_input)
            settings_form.addRow(self.party_override_label, self.party_override_input)
            settings_form.addRow("Stock parent group", self.stock_parent_input)
            settings_form.addRow("Godown", self.godown_input)
            settings_form.addRow("Batch", self.batch_input)
            settings_form.addRow("Output folder", self.output_dir_input)
            layout.addWidget(settings_box)

            actions = QHBoxLayout()
            preview_btn = QPushButton("Preview")
            preview_btn.clicked.connect(self._preview)
            generate_btn = QPushButton("Generate XML Bundle")
            generate_btn.clicked.connect(self._generate)
            out_btn = QPushButton("Choose Output Folder")
            out_btn.clicked.connect(self._pick_output_dir)
            back_btn = QPushButton("Back to Module")
            back_btn.clicked.connect(lambda: self.stack.setCurrentIndex(1))
            actions.addWidget(preview_btn)
            actions.addWidget(generate_btn)
            actions.addWidget(out_btn)
            actions.addStretch()
            actions.addWidget(back_btn)
            layout.addLayout(actions)

            self.summary = QTextEdit()
            self.summary.setReadOnly(True)
            self.summary.setPlaceholderText("Preview and generation summary will appear here.")
            layout.addWidget(self.summary, 1)
            return page

        def _open_module(self, module_name: str):
            self.current_module = module_name
            self.workspace_title.setText(f"{COMPANY_PROFILE_NAME} - {module_name} Entry")
            if module_name == "Sales":
                self.ledger_label.setText("Sales ledger")
                self.ledger_input.setText(DEFAULT_SALES_LEDGER)
                self.purchase_local_ledger_label.hide()
                self.purchase_local_ledger_input.hide()
                self.party_parent_label.setText("Party parent group")
                self.party_parent_input.setText(DEFAULT_SALES_PARTY_PARENT)
                self.party_override_label.setText("Party ledger override")
                self.party_override_input.clear()
                self.party_override_input.setPlaceholderText("Optional")
            else:
                self.ledger_label.setText("Interstate purchase ledger")
                self.ledger_input.setText(DEFAULT_PURCHASE_LEDGER)
                self.purchase_local_ledger_label.show()
                self.purchase_local_ledger_input.show()
                self.purchase_local_ledger_input.setText(DEFAULT_PURCHASE_LOCAL_LEDGER)
                self.party_parent_label.setText("Supplier parent group")
                self.party_parent_input.setText(DEFAULT_PURCHASE_PARTY_PARENT)
                self.party_override_label.setText("Supplier ledger override")
                self.party_override_input.setPlaceholderText("Optional exact Tally supplier ledger name")
            self.report_excel_path = ""
            self.report_path_input.clear()
            self.summary.clear()
            self.stack.setCurrentIndex(2)

        def _pick_report_excel(self):
            path, _ = QFileDialog.getOpenFileName(
                self,
                f"Choose {self.current_module} Excel Report",
                str(Path.home()),
                "Excel Files (*.xlsx *.xlsm *.xls *.xlsb)",
            )
            if not path:
                return
            self.report_excel_path = path
            self.report_path_input.setText(path)

        def _pick_master_xml(self):
            path, _ = QFileDialog.getOpenFileName(
                self,
                "Choose Existing Master XML",
                str(Path.home()),
                "XML Files (*.xml)",
            )
            if not path:
                return
            self.master_xml_path = path
            self.master_path_input.setText(path)
            snapshot = base.load_master_snapshot(path)
            if snapshot.company_name and not self.company_input.text().strip():
                self.company_input.setText(snapshot.company_name)
            if snapshot.godowns and base.normalize_key(self.godown_input.text()) in snapshot.godowns:
                self.godown_input.setText(snapshot.godowns[base.normalize_key(self.godown_input.text())])

        def _pick_output_dir(self):
            path = QFileDialog.getExistingDirectory(
                self,
                "Choose Output Folder",
                self.output_dir_input.text().strip() or str(Path.cwd()),
            )
            if path:
                self.output_dir_input.setText(path)

        def _ask_output_run_dir(self, base_output_dir: str) -> Optional[str]:
            default_name = _sanitize_output_folder_name(Path(self.report_excel_path).stem)
            folder_name, accepted = QInputDialog.getText(
                self,
                "Name Output Folder",
                "Folder name for this XML bundle:",
                text=default_name,
            )
            if not accepted:
                return None
            folder_name = _sanitize_output_folder_name(folder_name)
            if not folder_name:
                QMessageBox.information(self, "Folder Name Required", "Enter a folder name for this XML bundle.")
                return None
            return str(Path(base_output_dir).expanduser() / folder_name)

        def _build_config(self) -> HubConfig:
            config = HubConfig(
                company_name=self.company_input.text().strip(),
                stock_parent_group=self.stock_parent_input.text().strip() or DEFAULT_STOCK_PARENT,
                godown_name=self.godown_input.text().strip() or DEFAULT_GODOWN,
                batch_name=self.batch_input.text().strip() or DEFAULT_BATCH,
                output_dir=self.output_dir_input.text().strip() or str(Path.cwd()),
            )
            if self.current_module == "Sales":
                config.sales_ledger_name = self.ledger_input.text().strip() or DEFAULT_SALES_LEDGER
                config.sales_party_parent = self.party_parent_input.text().strip() or DEFAULT_SALES_PARTY_PARENT
            else:
                config.purchase_ledger_name = self.ledger_input.text().strip() or DEFAULT_PURCHASE_LEDGER
                config.purchase_local_ledger_name = self.purchase_local_ledger_input.text().strip() or DEFAULT_PURCHASE_LOCAL_LEDGER
                config.purchase_party_parent = self.party_parent_input.text().strip() or DEFAULT_PURCHASE_PARTY_PARENT
                config.purchase_party_name_override = self.party_override_input.text().strip()
            return config

        def _preview(self):
            if not self.report_excel_path:
                QMessageBox.information(self, "Report Required", "Load the report file first.")
                return
            snapshot = base.load_master_snapshot(self.master_xml_path) if self.master_xml_path else base.MasterSnapshot()
            try:
                if self.current_module == "Sales":
                    vouchers, names, missing, skipped = base.parse_sales_report_excel(self.report_excel_path, snapshot)
                    role = "Party ledgers"
                else:
                    vouchers, names, missing, skipped = parse_purchase_report_excel(self.report_excel_path, snapshot)
                    role = "Supplier ledgers"
            except Exception as ex:
                self.summary.setPlainText(str(ex))
                return
            lines = [
                f"Module: {self.current_module}",
                f"Vouchers detected: {len(vouchers)}",
                f"Stock lines detected: {sum(len(v.lines) for v in vouchers)}",
                f"{role} in report: {len(names)}",
                f"Missing stock masters after matching: {len(missing)}",
            ]
            if skipped:
                lines.append("")
                lines.append("Skipped rows:")
                lines.extend(skipped[:12])
            self.summary.setPlainText("\n".join(lines))

        def _generate(self):
            if not self.report_excel_path:
                QMessageBox.information(self, "Report Required", "Load the report file first.")
                return
            config = self._build_config()
            run_output_dir = self._ask_output_run_dir(config.output_dir)
            if not run_output_dir:
                return
            config.output_dir = run_output_dir
            state_path = str(Path(config.output_dir) / STATE_FILE)
            try:
                if self.current_module == "Sales":
                    result = base.convert_sales_report(
                        excel_path=self.report_excel_path,
                        master_xml_path=self.master_xml_path or None,
                        config=base.ConversionConfig(
                            company_name=config.company_name,
                            sales_ledger_name=config.sales_ledger_name,
                            party_parent_group=config.sales_party_parent,
                            stock_parent_group=config.stock_parent_group,
                            godown_name=config.godown_name,
                            batch_name=config.batch_name,
                            output_dir=config.output_dir,
                        ),
                        state_path=state_path,
                    )
                    bundle_name = "sales_import_bundle.zip"
                else:
                    result = convert_purchase_report(
                        excel_path=self.report_excel_path,
                        master_xml_path=self.master_xml_path or None,
                        config=config,
                        state_path=state_path,
                    )
                    bundle_name = "purchase_import_bundle.zip"
            except Exception as ex:
                QMessageBox.warning(self, "Generation Failed", str(ex))
                return

            lines = [
                f"Module: {self.current_module}",
                f"Generated vouchers: {len(result.vouchers)}",
                f"New ledgers: {len(result.ledger_names)}",
                f"Missing stock masters: {len(result.missing_stock_items)}",
                "",
                "Files:",
            ]
            lines.extend(result.output_files)
            if result.skipped_rows:
                lines.append("")
                lines.append("Skipped rows:")
                lines.extend(result.skipped_rows[:12])
            self.summary.setPlainText("\n".join(lines))
            QMessageBox.information(self, "Bundle Generated", f"Generated {bundle_name} in:\n{config.output_dir}")

    app = QApplication([])
    app.setApplicationName(APP_TITLE)
    window = HubWindow()
    window.show()
    app.exec()


def run_cli(args: argparse.Namespace) -> int:
    module = (args.module or "sales").strip().lower()
    if module not in {"sales", "purchase"}:
        raise ValueError("module must be 'sales' or 'purchase'")
    if not args.report_excel:
        raise ValueError("--report-excel is required for CLI mode.")

    config = HubConfig(
        company_name=args.company or "",
        sales_ledger_name=args.sales_ledger or DEFAULT_SALES_LEDGER,
        purchase_ledger_name=args.purchase_ledger or DEFAULT_PURCHASE_LEDGER,
        purchase_local_ledger_name=args.purchase_local_ledger or DEFAULT_PURCHASE_LOCAL_LEDGER,
        sales_party_parent=args.sales_party_parent or DEFAULT_SALES_PARTY_PARENT,
        purchase_party_parent=args.purchase_party_parent or DEFAULT_PURCHASE_PARTY_PARENT,
        purchase_party_name_override=args.purchase_party_override or "",
        stock_parent_group=args.stock_parent or DEFAULT_STOCK_PARENT,
        godown_name=args.godown or DEFAULT_GODOWN,
        batch_name=args.batch or DEFAULT_BATCH,
        output_dir=args.output_dir or str(Path.cwd()),
    )
    state_path = args.state_path or str(Path(config.output_dir) / STATE_FILE)

    if module == "sales":
        result = base.convert_sales_report(
            excel_path=args.report_excel,
            master_xml_path=args.master_xml,
            config=base.ConversionConfig(
                company_name=config.company_name,
                sales_ledger_name=config.sales_ledger_name,
                party_parent_group=config.sales_party_parent,
                stock_parent_group=config.stock_parent_group,
                godown_name=config.godown_name,
                batch_name=config.batch_name,
                output_dir=config.output_dir,
            ),
            state_path=state_path,
        )
    else:
        result = convert_purchase_report(
            excel_path=args.report_excel,
            master_xml_path=args.master_xml,
            config=config,
            state_path=state_path,
        )

    print(f"Module: {module}")
    print(f"Generated vouchers: {len(result.vouchers)}")
    print(f"New ledgers: {len(result.ledger_names)}")
    print(f"Missing stock masters: {len(result.missing_stock_items)}")
    print("Output files:")
    for file_path in result.output_files:
        print(file_path)
    if result.skipped_rows:
        print("Skipped rows:")
        for row in result.skipped_rows[:12]:
            print(row)
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=APP_TITLE)
    parser.add_argument("--module", default="sales", help="sales or purchase")
    parser.add_argument("--report-excel", help="Path to Sales/Purchase Excel report")
    parser.add_argument("--master-xml", default=None, help="Optional path to current Tally master XML")
    parser.add_argument("--company", default="", help="Company name override")
    parser.add_argument("--sales-ledger", default=DEFAULT_SALES_LEDGER, help="Sales ledger name")
    parser.add_argument("--purchase-ledger", default=DEFAULT_PURCHASE_LEDGER, help="Interstate purchase ledger name")
    parser.add_argument("--purchase-local-ledger", default=DEFAULT_PURCHASE_LOCAL_LEDGER, help="Kerala/local purchase ledger name")
    parser.add_argument("--sales-party-parent", default=DEFAULT_SALES_PARTY_PARENT, help="Sales party parent group")
    parser.add_argument("--purchase-party-parent", default=DEFAULT_PURCHASE_PARTY_PARENT, help="Purchase party parent group")
    parser.add_argument("--purchase-party-override", default="", help="Exact supplier ledger name to use for all purchase vouchers")
    parser.add_argument("--stock-parent", default=DEFAULT_STOCK_PARENT, help="Stock parent group for missing stock masters")
    parser.add_argument("--godown", default=DEFAULT_GODOWN, help="Godown name")
    parser.add_argument("--batch", default=DEFAULT_BATCH, help="Batch name")
    parser.add_argument("--output-dir", default=str(Path.cwd()), help="Output folder")
    parser.add_argument("--state-path", default=None, help="State file path")
    parser.add_argument("--gui", action="store_true", help="Launch GUI")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    if args.gui or not args.report_excel:
        launch_gui()
        return 0
    return run_cli(args)


if __name__ == "__main__":
    raise SystemExit(main())
