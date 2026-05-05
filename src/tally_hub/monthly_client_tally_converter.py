import argparse
import json
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


APP_TITLE = "Monthly Client Tally Converter"
STATE_FILE = "monthly_client_tally_state.json"


def sanitize_xml_text(text: object) -> str:
    if text is None:
        return ""
    cleaned = str(text).strip()
    return re.sub(r"[^\u0009\u000A\u000D\u0020-\uD7FF\uE000-\uFFFD]", "", cleaned)


def normalize_key(text: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", sanitize_xml_text(text).lower())


def parse_amount(raw_amount: object) -> Optional[Decimal]:
    text = sanitize_xml_text(raw_amount).replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def normalize_date(raw_date: object) -> Optional[str]:
    cleaned = sanitize_xml_text(raw_date)
    if not cleaned:
        return None
    for fmt in ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d/%m/%y"]:
        try:
            return datetime.strptime(cleaned, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return None


def format_amount(value: Decimal, places: int = 2) -> str:
    quant = Decimal("1") if places == 0 else Decimal("1." + ("0" * places))
    return format(value.quantize(quant), f".{places}f")


def format_qty(value: Decimal, unit_name: str) -> str:
    return f" {format_amount(value, 2)} {sanitize_xml_text(unit_name)}"


def safe_indent(root: ET.Element) -> None:
    try:
        ET.indent(root, space="  ")
    except AttributeError:
        pass


@dataclass
class SalesLine:
    item_name: str
    resolved_item_name: str
    hsn: str
    qty: Decimal
    unit_name: str
    amount: Decimal


@dataclass
class SalesVoucher:
    voucher_no: str
    date: str
    party_name: str
    total_amount: Decimal
    lines: List[SalesLine] = field(default_factory=list)


@dataclass
class MasterSnapshot:
    company_name: str = ""
    stock_items: Dict[str, str] = field(default_factory=dict)
    ledgers: Dict[str, str] = field(default_factory=dict)
    units: Dict[str, str] = field(default_factory=dict)
    godowns: Dict[str, str] = field(default_factory=dict)


@dataclass
class ConversionConfig:
    company_name: str = ""
    sales_ledger_name: str = "LOCAL SALES"
    party_parent_group: str = "Sundry Debtors"
    stock_parent_group: str = "Primary"
    godown_name: str = "Main Location"
    batch_name: str = "Primary Batch"
    fallback_unit: str = "Bag"
    output_dir: str = "."


@dataclass
class ConversionResult:
    vouchers: List[SalesVoucher]
    ledger_names: List[str]
    missing_stock_items: List[Tuple[str, str, str]]
    output_files: List[str]
    skipped_rows: List[str]
    summary_path: str


def _read_state(state_path: Path) -> Dict[str, List[str]]:
    if not state_path.exists():
        return {"known_ledgers": [], "known_stock_items": []}
    try:
        data = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return {"known_ledgers": [], "known_stock_items": []}
    if not isinstance(data, dict):
        return {"known_ledgers": [], "known_stock_items": []}
    known_ledgers = data.get("known_ledgers", [])
    known_stock_items = data.get("known_stock_items", [])
    if not isinstance(known_ledgers, list) or not isinstance(known_stock_items, list):
        return {"known_ledgers": [], "known_stock_items": []}
    return {
        "known_ledgers": [sanitize_xml_text(x) for x in known_ledgers if sanitize_xml_text(x)],
        "known_stock_items": [sanitize_xml_text(x) for x in known_stock_items if sanitize_xml_text(x)],
    }


def _write_state(state_path: Path, known_ledgers: Iterable[str], known_stock_items: Iterable[str]) -> None:
    payload = {
        "known_ledgers": sorted({sanitize_xml_text(x) for x in known_ledgers if sanitize_xml_text(x)}),
        "known_stock_items": sorted({sanitize_xml_text(x) for x in known_stock_items if sanitize_xml_text(x)}),
    }
    state_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _read_text_with_fallbacks(path: Path) -> str:
    raw_bytes = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def load_master_snapshot(master_xml_path: Optional[str]) -> MasterSnapshot:
    snapshot = MasterSnapshot()
    if not master_xml_path:
        return snapshot

    path = Path(master_xml_path)
    if not path.exists():
        raise FileNotFoundError(f"Master XML not found: {path}")

    company_pattern = re.compile(r"<SVCURRENTCOMPANY>(.*?)</SVCURRENTCOMPANY>", re.IGNORECASE)
    stock_pattern = re.compile(r'<STOCKITEM\b[^>]*\bNAME="([^"]+)"', re.IGNORECASE)
    ledger_pattern = re.compile(r'<LEDGER\b[^>]*\bNAME="([^"]+)"', re.IGNORECASE)
    unit_pattern = re.compile(r'<UNIT\b[^>]*\bNAME="([^"]+)"', re.IGNORECASE)
    godown_pattern = re.compile(r'<GODOWN\b[^>]*\bNAME="([^"]+)"', re.IGNORECASE)

    text = _read_text_with_fallbacks(path)

    company_match = company_pattern.search(text)
    if company_match:
        snapshot.company_name = sanitize_xml_text(company_match.group(1))

    for name in stock_pattern.findall(text):
        clean = sanitize_xml_text(name)
        if clean:
            snapshot.stock_items[normalize_key(clean)] = clean
    for name in ledger_pattern.findall(text):
        clean = sanitize_xml_text(name)
        if clean:
            snapshot.ledgers[normalize_key(clean)] = clean
    for name in unit_pattern.findall(text):
        clean = sanitize_xml_text(name)
        if clean:
            snapshot.units[normalize_key(clean)] = clean
    for name in godown_pattern.findall(text):
        clean = sanitize_xml_text(name)
        if clean:
            snapshot.godowns[normalize_key(clean)] = clean
    return snapshot


def detect_header_row(ws) -> int:
    required = {"Date", "Voucher No", "Party Name", "ItemName", "Amount", "Qty", "Unit"}
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        row_values = {sanitize_xml_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)}
        if required.issubset(row_values):
            return row_idx
    raise ValueError("Could not detect the header row in the Excel report.")


def parse_sales_report_excel(
    excel_path: str,
    master_snapshot: Optional[MasterSnapshot] = None,
) -> Tuple[List[SalesVoucher], List[str], List[Tuple[str, str, str]], List[str]]:
    master_snapshot = master_snapshot or MasterSnapshot()
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = detect_header_row(worksheet)
    header_map: Dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        header_name = sanitize_xml_text(worksheet.cell(header_row, col).value)
        if header_name:
            header_map[header_name] = col

    required_headers = ["Date", "Voucher No", "Party Name", "ItemName", "Amount", "Qty", "Unit"]
    missing_headers = [name for name in required_headers if name not in header_map]
    if missing_headers:
        raise ValueError("Missing required columns: " + ", ".join(missing_headers))

    grouped_rows = defaultdict(list)
    skipped_rows: List[str] = []
    item_resolution_cache: Dict[str, str] = {}
    missing_stock_candidates: Dict[str, Tuple[str, str, str]] = {}

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        voucher_no = sanitize_xml_text(worksheet.cell(row_idx, header_map["Voucher No"]).value)
        raw_date = worksheet.cell(row_idx, header_map["Date"]).value
        party_name = sanitize_xml_text(worksheet.cell(row_idx, header_map["Party Name"]).value)
        item_name = sanitize_xml_text(worksheet.cell(row_idx, header_map["ItemName"]).value)
        hsn = sanitize_xml_text(worksheet.cell(row_idx, header_map.get("HSN", 0)).value if "HSN" in header_map else "")
        qty = parse_amount(worksheet.cell(row_idx, header_map["Qty"]).value)
        unit_name = sanitize_xml_text(worksheet.cell(row_idx, header_map["Unit"]).value)
        amount_value = worksheet.cell(row_idx, header_map.get("Total", header_map["Amount"])).value
        amount = parse_amount(amount_value)

        if not any([voucher_no, raw_date, party_name, item_name, qty, amount]):
            continue
        if item_name.lower() == "total":
            continue

        normalized_date = normalize_date(raw_date)
        row_errors = []
        if not voucher_no:
            row_errors.append("voucher no")
        if not normalized_date:
            row_errors.append("date")
        if not party_name:
            row_errors.append("party")
        if not item_name:
            row_errors.append("item")
        if qty is None:
            row_errors.append("qty")
        if amount is None:
            row_errors.append("amount")
        if row_errors:
            skipped_rows.append(f"Row {row_idx}: {', '.join(row_errors)}")
            continue

        exact_item_name = item_resolution_cache.get(item_name)
        if not exact_item_name:
            normalized_item = normalize_key(item_name)
            exact_item_name = master_snapshot.stock_items.get(normalized_item, item_name)
            item_resolution_cache[item_name] = exact_item_name
        if exact_item_name == item_name and normalize_key(item_name) not in master_snapshot.stock_items:
            fallback_unit = unit_name or master_snapshot.units.get(normalize_key("Bag"), "Bag")
            missing_stock_candidates[normalize_key(item_name)] = (item_name, hsn, fallback_unit)

        resolved_unit = unit_name or master_snapshot.units.get(normalize_key("Bag"), "Bag")
        if unit_name and master_snapshot.units:
            resolved_unit = master_snapshot.units.get(normalize_key(unit_name), unit_name)

        grouped_rows[(normalized_date, voucher_no, party_name)].append(
            SalesLine(
                item_name=item_name,
                resolved_item_name=exact_item_name,
                hsn=hsn,
                qty=abs(qty) if qty is not None else Decimal("0"),
                unit_name=resolved_unit or "Bag",
                amount=abs(amount) if amount is not None else Decimal("0"),
            )
        )

    vouchers: List[SalesVoucher] = []
    for (voucher_date, voucher_no, party_name), lines in sorted(grouped_rows.items(), key=lambda item: (item[0][0], item[0][1])):
        total_amount = sum((line.amount for line in lines), Decimal("0"))
        vouchers.append(
            SalesVoucher(
                voucher_no=voucher_no,
                date=voucher_date,
                party_name=party_name,
                total_amount=total_amount,
                lines=lines,
            )
        )

    ledger_names = sorted({voucher.party_name for voucher in vouchers}, key=str.upper)
    missing_stock_items = [missing_stock_candidates[key] for key in sorted(missing_stock_candidates)]
    return vouchers, ledger_names, missing_stock_items, skipped_rows


def new_envelope(report_name: str, company_name: str = "") -> Tuple[ET.Element, ET.Element]:
    root = ET.Element("ENVELOPE")
    header = ET.SubElement(root, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"

    body = ET.SubElement(root, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")
    request_desc = ET.SubElement(import_data, "REQUESTDESC")
    ET.SubElement(request_desc, "REPORTNAME").text = report_name

    company_name = sanitize_xml_text(company_name)
    if company_name:
        static_vars = ET.SubElement(request_desc, "STATICVARIABLES")
        ET.SubElement(static_vars, "SVCURRENTCOMPANY").text = company_name

    request_data = ET.SubElement(import_data, "REQUESTDATA")
    return root, request_data


def _add_common_empty_lists(parent: ET.Element, tag_names: Sequence[str]) -> None:
    for tag_name in tag_names:
        ET.SubElement(parent, tag_name)


def build_sales_vouchers_xml(vouchers: Sequence[SalesVoucher], config: ConversionConfig, output_path: str) -> str:
    root, request_data = new_envelope("Vouchers", config.company_name)

    for voucher_row in vouchers:
        tally_message = ET.SubElement(request_data, "TALLYMESSAGE", attrib={"xmlns:UDF": "TallyUDF"})
        voucher = ET.SubElement(
            tally_message,
            "VOUCHER",
            attrib={"VCHTYPE": "Sales", "ACTION": "Create", "OBJVIEW": "Accounting Voucher View"},
        )
        ET.SubElement(voucher, "DATE").text = voucher_row.date
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = "Sales"
        ET.SubElement(voucher, "PARTYLEDGERNAME").text = sanitize_xml_text(voucher_row.party_name)
        ET.SubElement(voucher, "VOUCHERNUMBER").text = sanitize_xml_text(voucher_row.voucher_no)
        ET.SubElement(voucher, "PERSISTEDVIEW").text = "Accounting Voucher View"
        ET.SubElement(voucher, "ISINVOICE").text = "No"
        ET.SubElement(voucher, "EFFECTIVEDATE").text = voucher_row.date

        party_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(party_entry, "LEDGERNAME").text = sanitize_xml_text(voucher_row.party_name)
        ET.SubElement(party_entry, "ISDEEMEDPOSITIVE").text = "Yes"
        ET.SubElement(party_entry, "LEDGERFROMITEM").text = "No"
        ET.SubElement(party_entry, "ISPARTYLEDGER").text = "Yes"
        ET.SubElement(party_entry, "AMOUNT").text = format_amount(-voucher_row.total_amount)
        _add_common_empty_lists(
            party_entry,
            [
                "BANKALLOCATIONS.LIST",
                "BILLALLOCATIONS.LIST",
                "INTERESTCOLLECTION.LIST",
                "OLDAUDITENTRIES.LIST",
                "ACCOUNTAUDITENTRIES.LIST",
                "AUDITENTRIES.LIST",
                "INPUTCRALLOCS.LIST",
                "DUTYHEADDETAILS.LIST",
                "RATEDETAILS.LIST",
                "SUMMARYALLOCS.LIST",
                "TAXBILLALLOCATIONS.LIST",
                "TAXOBJECTALLOCATIONS.LIST",
                "VATSTATUTORYDETAILS.LIST",
            ],
        )

        sales_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(sales_entry, "LEDGERNAME").text = sanitize_xml_text(config.sales_ledger_name)
        ET.SubElement(sales_entry, "ISDEEMEDPOSITIVE").text = "No"
        ET.SubElement(sales_entry, "LEDGERFROMITEM").text = "No"
        ET.SubElement(sales_entry, "ISPARTYLEDGER").text = "No"
        ET.SubElement(sales_entry, "AMOUNT").text = format_amount(voucher_row.total_amount)

        for line in voucher_row.lines:
            inventory = ET.SubElement(sales_entry, "INVENTORYALLOCATIONS.LIST")
            ET.SubElement(inventory, "STOCKITEMNAME").text = sanitize_xml_text(line.resolved_item_name)
            ET.SubElement(inventory, "GSTSOURCETYPE").text = "Stock Item"
            ET.SubElement(inventory, "GSTITEMSOURCE").text = sanitize_xml_text(line.resolved_item_name)
            ET.SubElement(inventory, "HSNSOURCETYPE").text = "Stock Item"
            ET.SubElement(inventory, "HSNITEMSOURCE").text = sanitize_xml_text(line.resolved_item_name)
            ET.SubElement(inventory, "GSTOVRDNTYPEOFSUPPLY").text = "Goods"
            ET.SubElement(inventory, "GSTRATEINFERAPPLICABILITY").text = "As per Masters/Company"
            if line.hsn:
                ET.SubElement(inventory, "GSTHSNNAME").text = sanitize_xml_text(line.hsn)
            ET.SubElement(inventory, "GSTHSNINFERAPPLICABILITY").text = "As per Masters/Company"
            ET.SubElement(inventory, "ISDEEMEDPOSITIVE").text = "No"
            ET.SubElement(inventory, "AMOUNT").text = format_amount(line.amount)
            ET.SubElement(inventory, "ACTUALQTY").text = format_qty(line.qty, line.unit_name)
            ET.SubElement(inventory, "BILLEDQTY").text = format_qty(line.qty, line.unit_name)

            batch = ET.SubElement(inventory, "BATCHALLOCATIONS.LIST")
            ET.SubElement(batch, "GODOWNNAME").text = sanitize_xml_text(config.godown_name)
            ET.SubElement(batch, "BATCHNAME").text = sanitize_xml_text(config.batch_name)
            ET.SubElement(batch, "AMOUNT").text = format_amount(line.amount)
            ET.SubElement(batch, "ACTUALQTY").text = format_qty(line.qty, line.unit_name)
            ET.SubElement(batch, "BILLEDQTY").text = format_qty(line.qty, line.unit_name)
            _add_common_empty_lists(batch, ["ADDITIONALDETAILS.LIST", "VOUCHERCOMPONENTLIST.LIST"])

            _add_common_empty_lists(
                inventory,
                [
                    "DUTYHEADDETAILS.LIST",
                    "RATEDETAILS.LIST",
                    "SUPPLEMENTARYDUTYHEADDETAILS.LIST",
                    "TAXOBJECTALLOCATIONS.LIST",
                    "COSTTRACKALLOCATIONS.LIST",
                    "REFVOUCHERDETAILS.LIST",
                    "EXCISEALLOCATIONS.LIST",
                    "EXPENSEALLOCATIONS.LIST",
                ],
            )

        _add_common_empty_lists(
            sales_entry,
            [
                "BANKALLOCATIONS.LIST",
                "BILLALLOCATIONS.LIST",
                "INTERESTCOLLECTION.LIST",
                "OLDAUDITENTRIES.LIST",
                "ACCOUNTAUDITENTRIES.LIST",
                "AUDITENTRIES.LIST",
                "INPUTCRALLOCS.LIST",
                "DUTYHEADDETAILS.LIST",
                "RATEDETAILS.LIST",
                "SUMMARYALLOCS.LIST",
                "TAXBILLALLOCATIONS.LIST",
                "TAXOBJECTALLOCATIONS.LIST",
                "VATSTATUTORYDETAILS.LIST",
            ],
        )

    safe_indent(root)
    ET.ElementTree(root).write(output_path, encoding="utf-8", xml_declaration=True)
    return output_path


def build_ledger_masters_xml(ledger_names: Sequence[str], config: ConversionConfig, output_path: str) -> str:
    root, request_data = new_envelope("All Masters", config.company_name)
    for ledger_name in ledger_names:
        tally_message = ET.SubElement(request_data, "TALLYMESSAGE", attrib={"xmlns:UDF": "TallyUDF"})
        ledger = ET.SubElement(
            tally_message,
            "LEDGER",
            attrib={"NAME": sanitize_xml_text(ledger_name), "ACTION": "Create"},
        )
        name_list = ET.SubElement(ledger, "NAME.LIST", attrib={"TYPE": "String"})
        ET.SubElement(name_list, "NAME").text = sanitize_xml_text(ledger_name)
        ET.SubElement(ledger, "PARENT").text = sanitize_xml_text(config.party_parent_group)
        ET.SubElement(ledger, "OPENINGBALANCE").text = "0.00"
    safe_indent(root)
    ET.ElementTree(root).write(output_path, encoding="utf-8", xml_declaration=True)
    return output_path


def build_stock_masters_xml(
    missing_stock_items: Sequence[Tuple[str, str, str]],
    config: ConversionConfig,
    output_path: str,
) -> Optional[str]:
    if not missing_stock_items:
        return None

    root, request_data = new_envelope("All Masters", config.company_name)
    applicable_from = datetime.now().strftime("%Y%m%d")
    for item_name, hsn, unit_name in missing_stock_items:
        clean_item_name = sanitize_xml_text(item_name)
        clean_unit_name = sanitize_xml_text(unit_name or config.fallback_unit or "Bag")
        tally_message = ET.SubElement(request_data, "TALLYMESSAGE", attrib={"xmlns:UDF": "TallyUDF"})
        stock = ET.SubElement(
            tally_message,
            "STOCKITEM",
            attrib={"NAME": clean_item_name, "ACTION": "Create"},
        )
        name_list = ET.SubElement(stock, "NAME.LIST", attrib={"TYPE": "String"})
        ET.SubElement(name_list, "NAME").text = clean_item_name
        ET.SubElement(stock, "PARENT").text = sanitize_xml_text(config.stock_parent_group)
        ET.SubElement(stock, "GSTAPPLICABLE").text = "Applicable"
        ET.SubElement(stock, "GSTTYPEOFSUPPLY").text = "Goods"
        ET.SubElement(stock, "BASEUNITS").text = clean_unit_name
        ET.SubElement(stock, "VATBASEUNIT").text = clean_unit_name
        ET.SubElement(stock, "ISBATCHWISEON").text = "No"
        ET.SubElement(stock, "OPENINGBALANCE").text = format_qty(Decimal("0"), clean_unit_name)
        ET.SubElement(stock, "OPENINGVALUE").text = "0.00"

        gst_details = ET.SubElement(stock, "GSTDETAILS.LIST")
        ET.SubElement(gst_details, "APPLICABLEFROM").text = applicable_from
        ET.SubElement(gst_details, "TAXABILITY").text = "Nil Rated"
        ET.SubElement(gst_details, "SRCOFGSTDETAILS").text = "Specify Details Here"
        ET.SubElement(gst_details, "GSTCALCSLABONMRP").text = "No"
        state_details = ET.SubElement(gst_details, "STATEWISEDETAILS.LIST")
        ET.SubElement(state_details, "STATENAME").text = "Any"

        if hsn:
            hsn_details = ET.SubElement(stock, "HSNDETAILS.LIST")
            ET.SubElement(hsn_details, "APPLICABLEFROM").text = applicable_from
            ET.SubElement(hsn_details, "HSNCODE").text = sanitize_xml_text(hsn)
            ET.SubElement(hsn_details, "SRCOFHSNDETAILS").text = "Specify Details Here"

    safe_indent(root)
    ET.ElementTree(root).write(output_path, encoding="utf-8", xml_declaration=True)
    return output_path


def write_summary_file(
    output_dir: Path,
    vouchers: Sequence[SalesVoucher],
    new_ledgers: Sequence[str],
    missing_stock_items: Sequence[Tuple[str, str, str]],
    skipped_rows: Sequence[str],
) -> str:
    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "voucher_count": len(vouchers),
        "voucher_lines": sum(len(v.lines) for v in vouchers),
        "new_ledgers": list(new_ledgers),
        "missing_stock_items": [
            {"item_name": item_name, "hsn": hsn, "unit_name": unit_name}
            for item_name, hsn, unit_name in missing_stock_items
        ],
        "skipped_rows": list(skipped_rows),
    }
    summary_path = output_dir / "sales_conversion_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return str(summary_path)


def convert_sales_report(
    excel_path: str,
    master_xml_path: Optional[str],
    config: ConversionConfig,
    state_path: Optional[str] = None,
) -> ConversionResult:
    output_dir = Path(config.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    state_path_obj = Path(state_path or (output_dir / STATE_FILE))
    state = _read_state(state_path_obj)
    master_snapshot = load_master_snapshot(master_xml_path)

    if not config.company_name and master_snapshot.company_name:
        config.company_name = master_snapshot.company_name
    if master_snapshot.godowns and normalize_key(config.godown_name) in master_snapshot.godowns:
        config.godown_name = master_snapshot.godowns[normalize_key(config.godown_name)]
    if master_snapshot.units and normalize_key(config.fallback_unit) in master_snapshot.units:
        config.fallback_unit = master_snapshot.units[normalize_key(config.fallback_unit)]

    vouchers, ledger_names, missing_stock_items, skipped_rows = parse_sales_report_excel(excel_path, master_snapshot)

    known_ledgers = {sanitize_xml_text(name) for name in state["known_ledgers"]}
    known_ledgers.update(master_snapshot.ledgers.values())
    reserved_ledgers = {"cash"}
    new_ledgers = [
        name
        for name in ledger_names
        if sanitize_xml_text(name) not in known_ledgers and normalize_key(name) not in reserved_ledgers
    ]

    output_files: List[str] = []
    vouchers_path = output_dir / "sales_vouchers.xml"
    output_files.append(build_sales_vouchers_xml(vouchers, config, str(vouchers_path)))

    ledger_xml_path = output_dir / "party_ledgers.xml"
    output_files.append(build_ledger_masters_xml(new_ledgers, config, str(ledger_xml_path)))

    stock_xml_path = output_dir / "stock_masters_missing.xml"
    stock_path = build_stock_masters_xml(missing_stock_items, config, str(stock_xml_path))
    if stock_path:
        output_files.append(stock_path)
    elif stock_xml_path.exists():
        stock_xml_path.unlink()

    summary_path = write_summary_file(output_dir, vouchers, new_ledgers, missing_stock_items, skipped_rows)
    output_files.append(summary_path)

    bundle_path = output_dir / "sales_import_bundle.zip"
    with zipfile.ZipFile(bundle_path, "w", compression=zipfile.ZIP_DEFLATED) as bundle:
        for file_path in output_files:
            bundle.write(file_path, arcname=Path(file_path).name)
    output_files.append(str(bundle_path))

    updated_ledgers = known_ledgers | {sanitize_xml_text(name) for name in ledger_names}
    updated_stock_items = {sanitize_xml_text(name) for name in state["known_stock_items"]}
    updated_stock_items.update(master_snapshot.stock_items.values())
    for voucher in vouchers:
        for line in voucher.lines:
            updated_stock_items.add(sanitize_xml_text(line.resolved_item_name))
    _write_state(state_path_obj, updated_ledgers, updated_stock_items)

    return ConversionResult(
        vouchers=vouchers,
        ledger_names=new_ledgers,
        missing_stock_items=missing_stock_items,
        output_files=output_files,
        skipped_rows=skipped_rows,
        summary_path=summary_path,
    )


def launch_gui() -> None:
    try:
        from PySide6.QtWidgets import (
            QApplication,
            QFileDialog,
            QFormLayout,
            QGridLayout,
            QGroupBox,
            QHBoxLayout,
            QLabel,
            QLineEdit,
            QMainWindow,
            QMessageBox,
            QPushButton,
            QTextEdit,
            QVBoxLayout,
            QWidget,
        )
    except Exception as ex:
        raise RuntimeError("PySide6 is required for GUI mode.") from ex

    class ConverterWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle(APP_TITLE)
            self.resize(1080, 760)
            self.sales_excel_path = ""
            self.master_xml_path = ""
            self._build_ui()

        def _build_ui(self):
            root = QWidget()
            self.setCentralWidget(root)
            layout = QVBoxLayout(root)
            layout.setContentsMargins(18, 18, 18, 18)
            layout.setSpacing(14)

            intro = QLabel(
                "Load monthly sales Excel and optional Master XML, then generate Tally Sales voucher XML, "
                "new party ledgers, and missing stock masters in one bundle."
            )
            intro.setWordWrap(True)
            layout.addWidget(intro)

            file_box = QGroupBox("Source Files")
            file_layout = QGridLayout(file_box)
            self.sales_path_input = QLineEdit()
            self.master_path_input = QLineEdit()
            self.sales_path_input.setReadOnly(True)
            self.master_path_input.setReadOnly(True)
            sales_btn = QPushButton("Load Sales Excel")
            sales_btn.clicked.connect(self._pick_sales_excel)
            master_btn = QPushButton("Load Master XML")
            master_btn.clicked.connect(self._pick_master_xml)
            file_layout.addWidget(QLabel("Sales report"), 0, 0)
            file_layout.addWidget(self.sales_path_input, 0, 1)
            file_layout.addWidget(sales_btn, 0, 2)
            file_layout.addWidget(QLabel("Existing master XML"), 1, 0)
            file_layout.addWidget(self.master_path_input, 1, 1)
            file_layout.addWidget(master_btn, 1, 2)
            layout.addWidget(file_box)

            settings_box = QGroupBox("Import Settings")
            settings_form = QFormLayout(settings_box)
            self.company_input = QLineEdit()
            self.sales_ledger_input = QLineEdit("LOCAL SALES")
            self.party_parent_input = QLineEdit("Sundry Debtors")
            self.stock_parent_input = QLineEdit("Primary")
            self.godown_input = QLineEdit("Main Location")
            self.batch_input = QLineEdit("Primary Batch")
            self.output_dir_input = QLineEdit(str(Path.cwd()))
            settings_form.addRow("Company name", self.company_input)
            settings_form.addRow("Sales ledger", self.sales_ledger_input)
            settings_form.addRow("Party parent group", self.party_parent_input)
            settings_form.addRow("Stock parent group", self.stock_parent_input)
            settings_form.addRow("Godown", self.godown_input)
            settings_form.addRow("Batch", self.batch_input)
            settings_form.addRow("Output folder", self.output_dir_input)
            layout.addWidget(settings_box)

            actions = QHBoxLayout()
            generate_btn = QPushButton("Generate Tally XML Bundle")
            generate_btn.clicked.connect(self._generate)
            choose_output_btn = QPushButton("Choose Output Folder")
            choose_output_btn.clicked.connect(self._pick_output_dir)
            actions.addWidget(generate_btn)
            actions.addWidget(choose_output_btn)
            actions.addStretch()
            layout.addLayout(actions)

            self.summary = QTextEdit()
            self.summary.setReadOnly(True)
            self.summary.setPlaceholderText("Preview and generation summary will appear here.")
            layout.addWidget(self.summary, 1)

        def _pick_sales_excel(self):
            path, _ = QFileDialog.getOpenFileName(
                self,
                "Choose Sales Excel",
                str(Path.home()),
                "Excel Files (*.xlsx *.xlsm *.xls *.xlsb)",
            )
            if not path:
                return
            self.sales_excel_path = path
            self.sales_path_input.setText(path)
            self._refresh_preview()

        def _pick_master_xml(self):
            path, _ = QFileDialog.getOpenFileName(
                self,
                "Choose Existing Master XML",
                str(Path.home()),
                "XML Files (*.xml)",
            )
            if not path:
                return
            self.master_xml_path = path
            self.master_path_input.setText(path)
            try:
                snapshot = load_master_snapshot(path)
            except Exception as ex:
                QMessageBox.warning(self, "Master XML Error", str(ex))
                return
            if snapshot.company_name and not self.company_input.text().strip():
                self.company_input.setText(snapshot.company_name)
            if snapshot.godowns and normalize_key(self.godown_input.text()) in snapshot.godowns:
                self.godown_input.setText(snapshot.godowns[normalize_key(self.godown_input.text())])
            self._refresh_preview()

        def _pick_output_dir(self):
            path = QFileDialog.getExistingDirectory(
                self,
                "Choose Output Folder",
                self.output_dir_input.text().strip() or str(Path.cwd()),
            )
            if path:
                self.output_dir_input.setText(path)

        def _build_config(self) -> ConversionConfig:
            return ConversionConfig(
                company_name=self.company_input.text().strip(),
                sales_ledger_name=self.sales_ledger_input.text().strip() or "LOCAL SALES",
                party_parent_group=self.party_parent_input.text().strip() or "Sundry Debtors",
                stock_parent_group=self.stock_parent_input.text().strip() or "Primary",
                godown_name=self.godown_input.text().strip() or "Main Location",
                batch_name=self.batch_input.text().strip() or "Primary Batch",
                fallback_unit="Bag",
                output_dir=self.output_dir_input.text().strip() or str(Path.cwd()),
            )

        def _refresh_preview(self):
            if not self.sales_excel_path:
                return
            try:
                snapshot = load_master_snapshot(self.master_xml_path) if self.master_xml_path else MasterSnapshot()
                vouchers, ledger_names, missing_stock_items, skipped_rows = parse_sales_report_excel(
                    self.sales_excel_path,
                    snapshot,
                )
            except Exception as ex:
                self.summary.setPlainText(str(ex))
                return

            lines = [
                f"Vouchers detected: {len(vouchers)}",
                f"Stock lines detected: {sum(len(v.lines) for v in vouchers)}",
                f"Party ledgers in report: {len(ledger_names)}",
                f"Missing stock masters after matching: {len(missing_stock_items)}",
            ]
            if vouchers:
                lines.append("")
                lines.append("First vouchers:")
                for voucher in vouchers[:8]:
                    lines.append(
                        f"{voucher.date} | {voucher.voucher_no} | {voucher.party_name} | "
                        f"lines={len(voucher.lines)} | amount={format_amount(voucher.total_amount)}"
                    )
            if missing_stock_items:
                lines.append("")
                lines.append("Missing stock items:")
                for item_name, hsn, unit_name in missing_stock_items[:8]:
                    lines.append(f"{item_name} | HSN={hsn or '-'} | Unit={unit_name or '-'}")
            if skipped_rows:
                lines.append("")
                lines.append("Skipped rows:")
                lines.extend(skipped_rows[:10])
            self.summary.setPlainText("\n".join(lines))

        def _generate(self):
            if not self.sales_excel_path:
                QMessageBox.information(self, "Sales Report Required", "Load the monthly sales Excel file first.")
                return
            config = self._build_config()
            try:
                result = convert_sales_report(
                    excel_path=self.sales_excel_path,
                    master_xml_path=self.master_xml_path or None,
                    config=config,
                )
            except Exception as ex:
                QMessageBox.warning(self, "Generation Failed", str(ex))
                return

            lines = [
                f"Generated {len(result.vouchers)} sales vouchers.",
                f"New party ledgers XML entries: {len(result.ledger_names)}",
                f"Missing stock masters XML entries: {len(result.missing_stock_items)}",
                "",
                "Files:",
            ]
            lines.extend(result.output_files)
            if result.skipped_rows:
                lines.append("")
                lines.append("Skipped rows:")
                lines.extend(result.skipped_rows[:12])
            self.summary.setPlainText("\n".join(lines))
            QMessageBox.information(
                self,
                "Bundle Generated",
                f"Created Tally import files in:\n{config.output_dir}\n\nBundle: sales_import_bundle.zip",
            )

    app = QApplication([])
    app.setApplicationName(APP_TITLE)
    window = ConverterWindow()
    window.show()
    app.exec()


def run_cli(args: argparse.Namespace) -> int:
    config = ConversionConfig(
        company_name=args.company or "",
        sales_ledger_name=args.sales_ledger or "LOCAL SALES",
        party_parent_group=args.party_parent or "Sundry Debtors",
        stock_parent_group=args.stock_parent or "Primary",
        godown_name=args.godown or "Main Location",
        batch_name=args.batch or "Primary Batch",
        fallback_unit="Bag",
        output_dir=args.output_dir or str(Path.cwd()),
    )

    result = convert_sales_report(
        excel_path=args.sales_excel,
        master_xml_path=args.master_xml,
        config=config,
        state_path=args.state_path,
    )

    print(f"Generated vouchers: {len(result.vouchers)}")
    print(f"New party ledgers: {len(result.ledger_names)}")
    print(f"Missing stock masters: {len(result.missing_stock_items)}")
    print("Output files:")
    for file_path in result.output_files:
        print(file_path)
    if result.skipped_rows:
        print("Skipped rows:")
        for row in result.skipped_rows[:12]:
            print(row)
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=APP_TITLE)
    parser.add_argument("--sales-excel", help="Path to the monthly sales Excel report")
    parser.add_argument("--master-xml", help="Optional path to current Tally master XML", default=None)
    parser.add_argument("--company", help="Company name override", default="")
    parser.add_argument("--sales-ledger", help="Sales ledger name", default="LOCAL SALES")
    parser.add_argument("--party-parent", help="Parent group for new party ledgers", default="Sundry Debtors")
    parser.add_argument("--stock-parent", help="Parent group for missing stock items", default="Primary")
    parser.add_argument("--godown", help="Godown name for inventory allocations", default="Main Location")
    parser.add_argument("--batch", help="Batch name for inventory allocations", default="Primary Batch")
    parser.add_argument("--output-dir", help="Output folder", default=str(Path.cwd()))
    parser.add_argument("--state-path", help="Optional state JSON path", default=None)
    parser.add_argument("--gui", action="store_true", help="Launch the desktop app")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    if args.gui or not args.sales_excel:
        launch_gui()
        return 0
    return run_cli(args)


if __name__ == "__main__":
    raise SystemExit(main())
