import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_template(output_path):
    """
    Generates a blank Excel template for the user.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Totals"
    
    headers = ["Date", "Total Amount"]
    ws.append(headers)
    
    # Apply some basic styling to headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-fit columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20

    wb.save(output_path)

def read_input_excel(file_path):
    """
    Reads the input Excel file and returns a list of dictionaries:
    [{'Date': '01-03-2018', 'Total Amount': 183000}, ...]
    """
    df = pd.read_excel(file_path)
    
    # Basic validation
    if 'Date' not in df.columns or 'Total Amount' not in df.columns:
        raise ValueError("Input file must contain 'Date' and 'Total Amount' columns.")
    
    # Drop rows where Date or Total Amount is missing
    df = df.dropna(subset=['Date', 'Total Amount'])
    
    # Ensure Date is string format
    df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d-%m-%Y')
    df['Total Amount'] = pd.to_numeric(df['Total Amount'], errors='coerce')
    df = df.dropna(subset=['Total Amount'])
    
    return df.to_dict('records')

def write_output_excel(data_rows, output_path):
    """
    Writes the generated rows to an Excel file with hidden technical columns.
    data_rows format:
    [{'Date': '...', 'Voucher No': '...', 'Particulars': '...', 'Amount': 5420, 
      '_BatchID': '...', '_Seed': '...', '_Mode': '...'}, ...]
    """
    if not data_rows:
        return

    df = pd.DataFrame(data_rows)
    
    # Reorder columns to ensure public ones are first, technical ones last
    public_cols = ['Date', 'Voucher No', 'Particulars', 'Amount', 'Narration']
    tech_cols = [c for c in df.columns if c not in public_cols]
    df = df[public_cols + tech_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconstructed Purchases"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Styling and column adjustments
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Freeze the top row
    ws.freeze_panes = "A2"

    # Set column widths
    ws.column_dimensions['A'].width = 15 # Date
    ws.column_dimensions['B'].width = 15 # Voucher No
    ws.column_dimensions['C'].width = 30 # Particulars
    ws.column_dimensions['D'].width = 15 # Amount
    ws.column_dimensions['E'].width = 30 # Narration

    # Hide technical columns
    for i, col_name in enumerate(df.columns, start=1):
        if col_name in tech_cols:
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].hidden = True

    # Add a totals row at the bottom for Amount
    last_row = ws.max_row
    ws.cell(row=last_row + 1, column=3, value="TOTAL").font = Font(bold=True)
    
    # Formula to sum the amount column (D)
    sum_formula = f"=SUM(D2:D{last_row})"
    total_cell = ws.cell(row=last_row + 1, column=4, value=sum_formula)
    total_cell.font = Font(bold=True)

    wb.save(output_path)

def read_final_excel(filepath):
    """
    Reads a manually finalized "Reconstructed Purchases" Excel file 
    and returns a list of DayReconstruction-like objects suitable for TallyXMLEngine.
    """
    df = pd.read_excel(filepath)
    
    # Validation
    required = ['Date', 'Amount']
    for req in required:
        if req not in df.columns:
            raise ValueError(f"Imported Excel must contain '{req}' column.")
            
    # Drop totals row if present
    if 'Particulars' in df.columns:
        df = df[df['Particulars'] != "TOTAL"]
        
    df = df.dropna(subset=['Date', 'Amount'])
    
    # Ensure formats
    df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d-%m-%Y')
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
    df = df.dropna(subset=['Amount'])
    
    # Clean NaNs
    df = df.fillna("")
    
    days_dict = {}
    from core.project_manager import DayReconstruction
    
    for _, row in df.iterrows():
        date_str = row['Date']
        if date_str not in days_dict:
            days_dict[date_str] = DayReconstruction(date_str, 0) # total doesn't matter for XML
            
        entry = {
            'Amount': row['Amount'],
            'Voucher No': str(row.get('Voucher No', '')).strip(),
            'Particulars': str(row.get('Particulars', '')).strip(),
            'Narration': str(row.get('Narration', '')).strip()
        }
        days_dict[date_str].entries.append(entry)
        
    return list(days_dict.values())
